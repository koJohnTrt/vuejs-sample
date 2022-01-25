from io import BytesIO
import os
from xml.etree.ElementTree import Comment
from openpyxl.styles.borders import Side
import pandas as pd
import json

from fastapi import status

from fastapi import UploadFile, File
from decimal import Decimal
from styleframe import StyleFrame, Styler
from openpyxl import load_workbook
from openpyxl.styles import Font, Border
from fastapi.responses import JSONResponse
from openpyxl.comments import Comment


import configs.messages as msg


def create_excel_data(data, filename, columns):
    dict_new = []
    for i in data:
        dict_data = []
        for j in i:
            if type(j) is Decimal:
                j = int(j)
            dict_data.append(j)
        dict_new.append(dict_data)
    df2 = pd.DataFrame(dict_new, columns=columns)
    sf = StyleFrame(df2)
    style_column = Styler(bg_color='#D9D9D9', font_size = 10, font_color='#000000', comment_text="this is column comment") 
    style_header = Styler(bg_color='#0070C0', font_size = 12, font_color='#FFFFFF', comment_text="this is header comment") 
    sf.apply_style_by_indexes(sf.index[:], style_column)
    sf.apply_headers_style(style_header, style_index_header=True, cols_to_style=None)

    sf.to_excel(filename, best_fit=columns).save()


def create_excel_from_template(data, filename, columns):
    template_file = "static/excel_templates/" + os.path.splitext(filename)[0] + "_template.xlsx"

    wb = pd.ExcelFile(template_file)
    sheets = wb.sheet_names

    df = pd.DataFrame(data, columns=columns)

    sf = StyleFrame.read_excel_as_template(template_file, df, read_comments=True)
    writer = pd.ExcelWriter(filename)
    for sheet in sheets:
        sf.to_excel(writer, sheet_name=sheet)
    writer.save()


async def get_excel_data(excel_file: UploadFile = File(...)):
    if excel_file.filename == "":
        response = { 'code': "ME0092", 'message': msg.ME0092 }
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content=response
        )

    contents = await excel_file.read()
    wb = load_workbook(BytesIO(contents))
    sheets = wb.sheetnames
    json_data = []

    # フィルターがかかっている場合エラーにする
    for sheetname in sheets:
        ws = wb[sheetname]
        if len(ws.auto_filter.filterColumn) > 0:
            response = { 'code': "ME0093", 'message': msg.ME0093 }
            return JSONResponse(
                status_code=status.HTTP_400_BAD_REQUEST,
                content=response
            )

    for sheetname in sheets:
        df = pd.read_excel(contents, sheetname).fillna('')
        data = df.to_json(orient = 'records')
        data = eval(data)
        json_data += data

    return json_data


def create_excel_from_template_openpyxl(data, filename, columns):
    template_file = "static/excel_templates/" + os.path.splitext(filename)[0] + "_template.xlsx"

    wb = load_workbook(template_file)

    # Get first sheet
    ws = wb[wb.sheetnames[0]]

    # Define cell style
    font = Font(color="000000")
    bd = Side(border_style='thin')
    border = Border(left=bd, top=bd, right=bd, bottom=bd)

    # Set cell value, style
    x = 2
    for row in data:
        y = 1
        for item in row:
            ws.cell(row=x, column=y).value = item
            ws.cell(row=x, column=y).font = font
            ws.cell(row=x, column=y).border = border
            ws.cell(row=x, column=y).comment = Comment("this is " + str(item), 'Admin')
            y += 1
        x += 1

    # Save the file
    wb.save(filename)


async def get_excel_data_openpyxl(excel_file: UploadFile = File(...)):
    if excel_file.filename == "":
        response = { 'code': "ME0092", 'message': msg.ME0092 }
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content=response
        )

    contents = await excel_file.read()
    wb = load_workbook(BytesIO(contents))
    json_data = []
    header = []
    data = []

    ws = wb[wb.sheetnames[0]]
    max_col = ws.max_column

    # header
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_col):
        for cell in row:
            header.append(cell.value)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        max_row = ws.max_row

        # フィルターがかかっている場合エラーにする
        # if len(ws.auto_filter.filterColumn) > 0:
        #     response = { 'code': "ME0093", 'message': msg.ME0093 }
        #     return JSONResponse(
        #         status_code=status.HTTP_400_BAD_REQUEST,
        #         content=response
        #     )

        # data
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):

            # Get display data only
            if ws.row_dimensions[row[0].row].hidden == False:
                row_data = []
                for cell in row:
                    row_data.append(cell.value)

                data.append(row_data)

    # json data
    for row in data:
        params = {}
        for key, val in zip(header, row):
            params.update({key: val})
        json_data.append(params)

    return json_data